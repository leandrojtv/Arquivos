import csv
import importlib
import os
import sqlite3
import unicodedata
import uuid
from functools import wraps
from io import BytesIO, StringIO
from pathlib import Path

from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
LEGACY_DB = BASE_DIR / "people.db"
TERADATA_DRIVER_DIR = Path(
    os.environ.get("TERADATA_JDBC_DIR", BASE_DIR / "drivers" / "teradata")
).resolve()
if os.environ.get("DATABASE_PATH"):
    DB_PATH = Path(os.environ["DATABASE_PATH"])
elif LEGACY_DB.exists():
    DB_PATH = LEGACY_DB
else:
    DB_PATH = DATA_DIR / "people.db"

DB_PATH = DB_PATH.resolve()
TERADATA_DRIVER_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.secret_key = "change-me"  # Needed for flash messages.
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin")
IMPORT_CACHE = {}
DEFAULT_EXTRACTOR_GESTOR = {
    "name": "Gestor Padrão (Metadados)",
    "secretaria": "Extrações",
    "coordenacao": "Automático",
    "email": "gestor.meta@exemplo.gov",
}


def init_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS gestors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            secretaria TEXT NOT NULL,
            coordenacao TEXT NOT NULL,
            email TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS extraction_jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            connector TEXT NOT NULL,
            extraction_type TEXT NOT NULL,
            mode TEXT NOT NULL,
            host TEXT,
            jdbc_url TEXT,
            connection_type TEXT,
            database_name TEXT,
            password TEXT,
            username TEXT,
            extra_params TEXT,
            status TEXT NOT NULL DEFAULT 'pending',
            progress INTEGER NOT NULL DEFAULT 0,
            log TEXT,
            error TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS bases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            ambiente TEXT,
            descricao TEXT,
            gestor_id INTEGER NOT NULL,
            substituto1_id INTEGER,
            substituto2_id INTEGER,
            source_connector TEXT NOT NULL DEFAULT 'manual',
            source_job_id INTEGER,
            FOREIGN KEY (gestor_id) REFERENCES gestors(id),
            FOREIGN KEY (substituto1_id) REFERENCES gestors(id),
            FOREIGN KEY (substituto2_id) REFERENCES gestors(id),
            FOREIGN KEY (source_job_id) REFERENCES extraction_jobs(id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        INSERT INTO users (username, password)
        VALUES (?, ?)
        ON CONFLICT(username) DO UPDATE SET password=excluded.password
        """,
        (ADMIN_USERNAME, ADMIN_PASSWORD),
    )
    conn.commit()
    ensure_default_extractor_gestor()
    migrate_bases_nullable()
    migrate_bases_sources()
    migrate_extraction_jobs()
    conn.close()


def migrate_bases_nullable():
    conn = sqlite3.connect(DB_PATH)
    columns = conn.execute("PRAGMA table_info(bases)").fetchall()
    if not columns:
        conn.close()
        return

    notnull_map = {col[1]: bool(col[3]) for col in columns}
    needs_migration = any(
        [
            notnull_map.get("substituto1_id"),
            notnull_map.get("substituto2_id"),
            notnull_map.get("ambiente"),
            notnull_map.get("descricao"),
        ]
    )

    if not needs_migration:
        conn.close()
        return

    conn.execute("ALTER TABLE bases RENAME TO bases_old")
    conn.execute(
        """
        CREATE TABLE bases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            ambiente TEXT,
            descricao TEXT,
            gestor_id INTEGER NOT NULL,
            substituto1_id INTEGER,
            substituto2_id INTEGER,
            source_connector TEXT NOT NULL DEFAULT 'manual',
            source_job_id INTEGER,
            FOREIGN KEY (gestor_id) REFERENCES gestors(id),
            FOREIGN KEY (substituto1_id) REFERENCES gestors(id),
            FOREIGN KEY (substituto2_id) REFERENCES gestors(id),
            FOREIGN KEY (source_job_id) REFERENCES extraction_jobs(id)
        )
        """
    )
    conn.execute(
        """
        INSERT INTO bases (id, name, ambiente, descricao, gestor_id, substituto1_id, substituto2_id, source_connector, source_job_id)
        SELECT id, name, ambiente, descricao, gestor_id, substituto1_id, substituto2_id, 'manual', NULL FROM bases_old
        """
    )
    conn.execute("DROP TABLE bases_old")
    conn.commit()
    conn.close()


def migrate_bases_sources():
    conn = sqlite3.connect(DB_PATH)
    columns = conn.execute("PRAGMA table_info(bases)").fetchall()
    names = {col[1] for col in columns}
    if "source_connector" not in names:
        conn.execute("ALTER TABLE bases ADD COLUMN source_connector TEXT NOT NULL DEFAULT 'manual'")
    if "source_job_id" not in names:
        conn.execute("ALTER TABLE bases ADD COLUMN source_job_id INTEGER")
    conn.commit()
    conn.close()


def migrate_extraction_jobs():
    conn = sqlite3.connect(DB_PATH)
    columns = conn.execute("PRAGMA table_info(extraction_jobs)").fetchall()
    names = {col[1] for col in columns}
    if columns and "host" not in names:
        conn.execute("ALTER TABLE extraction_jobs ADD COLUMN host TEXT")
    if columns and "password" not in names:
        conn.execute("ALTER TABLE extraction_jobs ADD COLUMN password TEXT")
    conn.commit()
    conn.close()


def ensure_default_extractor_gestor():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    existing = conn.execute(
        "SELECT id FROM gestors WHERE email = ?", (DEFAULT_EXTRACTOR_GESTOR["email"],)
    ).fetchone()
    if existing:
        conn.close()
        return existing["id"]

    cursor = conn.execute(
        """
        INSERT INTO gestors (name, secretaria, coordenacao, email)
        VALUES (?, ?, ?, ?)
        """,
        (
            DEFAULT_EXTRACTOR_GESTOR["name"],
            DEFAULT_EXTRACTOR_GESTOR["secretaria"],
            DEFAULT_EXTRACTOR_GESTOR["coordenacao"],
            DEFAULT_EXTRACTOR_GESTOR["email"],
        ),
    )
    conn.commit()
    gid = cursor.lastrowid
    conn.close()
    return gid


def query_db(query, params=()):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.execute(query, params)
    rows = cur.fetchall()
    conn.commit()
    conn.close()
    return rows


def execute_db(query, params=()):
    conn = sqlite3.connect(DB_PATH)
    conn.execute(query, params)
    conn.commit()
    conn.close()


def load_optional_module(module_name):
    spec = importlib.util.find_spec(module_name)
    if not spec:
        return None
    return importlib.import_module(module_name)


def find_teradata_jars():
    return [jar for jar in TERADATA_DRIVER_DIR.glob("*.jar") if jar.is_file()]


def normalize_field(label: str) -> str:
    cleaned = (
        unicodedata.normalize("NFKD", label)
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
        .replace(" ", "")
        .replace("_", "")
    )
    return cleaned


def bulk_insert(records):
    conn = sqlite3.connect(DB_PATH)
    conn.executemany(
        "INSERT INTO gestors (name, secretaria, coordenacao, email) VALUES (?, ?, ?, ?)",
        records,
    )
    conn.commit()
    conn.close()


def bulk_insert_bases(records):
    conn = sqlite3.connect(DB_PATH)
    prepared = []
    for rec in records:
        if len(rec) == 6:
            prepared.append((*rec, "import", None))
        elif len(rec) == 7:
            prepared.append((*rec, None))
        else:
            prepared.append(rec)
    conn.executemany(
        """
        INSERT INTO bases (name, ambiente, descricao, gestor_id, substituto1_id, substituto2_id, source_connector, source_job_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        prepared,
    )
    conn.commit()
    conn.close()


def get_import_bucket():
    token = session.get("import_token")
    if not token:
        token = uuid.uuid4().hex
        session["import_token"] = token
    if token not in IMPORT_CACHE:
        IMPORT_CACHE[token] = {}
    return IMPORT_CACHE[token]


def get_flow_bucket(flow):
    bucket = get_import_bucket()
    if flow not in bucket:
        bucket[flow] = {}
    return bucket[flow]


def clear_import_state(flow=None):
    bucket = get_import_bucket()
    if flow:
        bucket.pop(flow, None)
    else:
        bucket.clear()


def create_extraction_job(connector, extraction_type, mode, config):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.execute(
        """
        INSERT INTO extraction_jobs (connector, extraction_type, mode, host, jdbc_url, connection_type, database_name, password, username, extra_params)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            connector,
            extraction_type,
            mode,
            config.get("host"),
            config.get("jdbc_url"),
            config.get("connection_type"),
            config.get("database_name"),
            config.get("password"),
            config.get("username"),
            config.get("extra_params"),
        ),
    )
    conn.commit()
    job_id = cur.lastrowid
    conn.close()
    return job_id


def update_extraction_job(job_id, status=None, progress=None, log=None, error=None):
    conn = sqlite3.connect(DB_PATH)
    sets = []
    params = []
    if status is not None:
        sets.append("status = ?")
        params.append(status)
    if progress is not None:
        sets.append("progress = ?")
        params.append(progress)
    if log is not None:
        sets.append("log = ?")
        params.append(log)
    if error is not None:
        sets.append("error = ?")
        params.append(error)
    sets.append("updated_at = CURRENT_TIMESTAMP")
    params.append(job_id)
    conn.execute(f"UPDATE extraction_jobs SET {', '.join(sets)} WHERE id = ?", params)
    conn.commit()
    conn.close()


def append_job_log(job_id, message, reset=False):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    existing = conn.execute("SELECT log FROM extraction_jobs WHERE id = ?", (job_id,)).fetchone()
    current_log = existing["log"] if existing and existing["log"] else ""
    new_log = message if reset or not current_log else f"{current_log}\n{message}"
    conn.execute(
        "UPDATE extraction_jobs SET log = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (new_log, job_id),
    )
    conn.commit()
    conn.close()


def get_job(job_id):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    job = conn.execute("SELECT * FROM extraction_jobs WHERE id = ?", (job_id,)).fetchone()
    conn.close()
    return job


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not session.get("user"):
            flash("Faça login para continuar.", "error")
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)

    return wrapper


def parse_csv(file_storage, delimiter):
    content = file_storage.stream.read().decode("utf-8-sig")
    reader = csv.DictReader(StringIO(content), delimiter=delimiter)
    records = []
    for row in reader:
        normalized = {normalize_field(k): (v or "").strip() for k, v in row.items()}
        name = normalized.get("gestor") or normalized.get("nome")
        secretaria = normalized.get("secretaria")
        coordenacao = normalized.get("coordenacao")
        email = normalized.get("email")
        if name and secretaria and coordenacao and email:
            records.append((name, secretaria, coordenacao, email))
    return records


def parse_xlsx(file_storage):
    file_bytes = BytesIO(file_storage.read())
    workbook = load_workbook(filename=file_bytes, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_field(str(h)) for h in rows[0] if h is not None]
    records = []
    for data_row in rows[1:]:
        values = [str(cell).strip() if cell is not None else "" for cell in data_row]
        row_dict = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        name = row_dict.get("gestor") or row_dict.get("nome")
        secretaria = row_dict.get("secretaria")
        coordenacao = row_dict.get("coordenacao")
        email = row_dict.get("email")
        if name and secretaria and coordenacao and email:
            records.append((name, secretaria, coordenacao, email))
    return records


def build_jdbc_url(host, database_name, connection_type, extra_params):
    if not host:
        return ""
    suffix = []
    if database_name:
        suffix.append(f"DATABASE={database_name}")
    if connection_type:
        suffix.append(f"LOGMECH={connection_type}")
    if extra_params:
        cleaned = extra_params.strip().strip(",")
        if cleaned:
            suffix.append(cleaned)
    formatted = ",".join(suffix)
    return f"jdbc:teradata://{host}/{formatted}" if formatted else f"jdbc:teradata://{host}"


def open_teradata_connection(config):
    jdbc_url = config.get("jdbc_url")
    username = config.get("username")
    password = config.get("password")
    if not jdbc_url or not username or not password:
        raise ValueError("Preencha JDBC, usuário e senha para conectar.")

    errors = []
    teradatasql = load_optional_module("teradatasql")
    if teradatasql:
        try:
            conn = teradatasql.connect(url=jdbc_url, user=username, password=password)
            return conn, "python"
        except Exception as exc:  # pragma: no cover - depende do driver
            errors.append(f"Driver Python falhou: {exc}")

    jdbc_jars = find_teradata_jars()
    jaydebeapi = load_optional_module("jaydebeapi")
    if not jdbc_jars:
        errors.append(f"Driver JDBC não encontrado em {TERADATA_DRIVER_DIR}")
    if not jaydebeapi:
        errors.append("Dependências JDBC ausentes (jaydebeapi/JPype1 não instaladas)")
    if jdbc_jars and jaydebeapi:
        try:
            conn = jaydebeapi.connect(
                "com.teradata.jdbc.TeraDriver",
                jdbc_url,
                {"user": username, "password": password},
                [str(j) for j in jdbc_jars],
            )
            return conn, "jdbc"
        except Exception as exc:  # pragma: no cover - depende do driver
            errors.append(f"Falha ao usar driver JDBC: {exc}")

    if not errors:
        errors.append("Driver Teradata não está disponível.")
    raise RuntimeError("; ".join(errors))


def test_teradata_connection(config):
    try:
        conn, _ = open_teradata_connection(config)
        conn.close()
        return True, "Conexão bem-sucedida."
    except Exception as exc:  # pragma: no cover - ambiente varia
        extra = ""
        if not find_teradata_jars():
            extra = f" Copie o JDBC para {TERADATA_DRIVER_DIR} e reinicie a aplicação."
        return False, f"Não foi possível conectar: {exc}.{extra}" if str(exc) else "Falha ao conectar com o driver JDBC."


def fetch_teradata_metadata(config):
    # Placeholder simples: tenta usar o driver se existir, caso contrário devolve amostra.
    database_name = config.get("database_name") or "database"
    sample_note = None
    try:
        conn, _ = open_teradata_connection(config)
        query = """
           select d.DatabaseName, d.CommentString
           FROM DBC.DatabasesV AS d
           where DBKind = 'D'
        """
        cur = conn.cursor()
        cur.execute(query)
        rows = cur.fetchall()
        conn.close()
        return [
            {"DatabaseName": row[0], "CommentString": row[1] if len(row) > 1 else ""}
            for row in rows
        ], None
    except Exception as exc:
        sample_note = (
            f"Extração simulada (driver JDBC indisponível nesta execução: {exc})."
            if str(exc)
            else "Extração simulada (driver JDBC indisponível nesta execução)."
        )

    return [
        {"DatabaseName": database_name, "CommentString": "Base importada via simulação."},
        {"DatabaseName": f"{database_name}_ANALYTICS", "CommentString": "Exemplo de metadado."},
    ], sample_note


def upsert_bases_from_metadata(rows, mode, job_id, gestor_id):
    imported = 0
    errors = []

    if mode == "full":
        execute_db("DELETE FROM bases WHERE source_connector = 'teradata'")

    for entry in rows:
        name = (entry.get("DatabaseName") or "").strip()
        descricao = (entry.get("CommentString") or "").strip() or None

        if not name:
            errors.append("Linha ignorada por falta do nome do database.")
            continue

        existing = query_db("SELECT id, source_connector FROM bases WHERE name = ?", (name,))
        if existing:
            record = existing[0]
            if record["source_connector"] not in (None, "teradata"):
                errors.append(f"Base '{name}' foi criada manualmente/importada e não será sobrescrita.")
                continue

            execute_db(
                """
                UPDATE bases
                SET descricao = ?, gestor_id = ?, source_connector = 'teradata', source_job_id = ?
                WHERE id = ?
                """,
                (descricao, gestor_id, job_id, record["id"]),
            )
        else:
            execute_db(
                """
                INSERT INTO bases (name, ambiente, descricao, gestor_id, substituto1_id, substituto2_id, source_connector, source_job_id)
                VALUES (?, ?, ?, ?, ?, ?, 'teradata', ?)
                """,
                (name, None, descricao, gestor_id, None, None, job_id),
            )
        imported += 1

    return imported, errors


def run_teradata_job(config, mode, extraction_type, job_id):
    append_job_log(job_id, "Iniciando extração de metadados do Teradata...", reset=True)
    update_extraction_job(job_id, status="running", progress=10, error=None)

    rows, note = fetch_teradata_metadata(config)
    if note:
        append_job_log(job_id, note)

    append_job_log(job_id, "Processando bases extraídas...")
    update_extraction_job(job_id, progress=40)

    gestor_id = ensure_default_extractor_gestor()
    imported, errors = upsert_bases_from_metadata(rows, mode, job_id, gestor_id)

    progress = 100 if rows else 0
    status = "success" if not errors else "completed"
    error_text = "\n".join(errors) if errors else None

    update_extraction_job(job_id, status=status, progress=progress, error=error_text)
    append_job_log(job_id, f"Linhas recebidas: {len(rows)}. Bases aplicadas: {imported}.")
    if errors:
        append_job_log(job_id, "Ocorreram avisos durante a execução:")
        for err in errors:
            append_job_log(job_id, f"- {err}")

    return {
        "total": len(rows),
        "imported": imported,
        "errors": errors,
        "note": note,
    }


def parse_tabular(file_storage, delimiter):
    filename = file_storage.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower()

    if ext == "csv":
        content = file_storage.stream.read().decode("utf-8-sig")
        reader = csv.DictReader(StringIO(content), delimiter=delimiter or ";")
        headers = [h or "Coluna" for h in (reader.fieldnames or [])]
        rows = []
        for row in reader:
            rows.append({h: (row.get(h, "") or "").strip() for h in headers})
        return headers, rows

    if ext in {"xlsx", "xls"}:
        file_bytes = BytesIO(file_storage.read())
        workbook = load_workbook(filename=file_bytes, data_only=True)
        sheet = workbook.active
        excel_rows = list(sheet.iter_rows(values_only=True))
        if not excel_rows:
            return [], []
        headers = [str(h) if h is not None else "Coluna" for h in excel_rows[0]]
        rows = []
        for data_row in excel_rows[1:]:
            values = [str(cell).strip() if cell is not None else "" for cell in data_row]
            mapped = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            rows.append(mapped)
        return headers, rows

    raise ValueError("Formato não suportado")


@app.route("/")
@login_required
def landing():
    total_bases = query_db("SELECT COUNT(*) as total FROM bases")[0]["total"]
    total_gestors = query_db("SELECT COUNT(*) as total FROM gestors")[0]["total"]
    with_gestor = query_db("SELECT COUNT(*) as total FROM bases WHERE gestor_id IS NOT NULL")[0]["total"]
    without_gestor = max(total_bases - with_gestor, 0)
    coverage_percent = round((with_gestor / total_bases) * 100, 1) if total_bases else 0

    return render_template(
        "landing.html",
        total_bases=total_bases,
        total_gestors=total_gestors,
        with_gestor=with_gestor,
        without_gestor=without_gestor,
        coverage_percent=coverage_percent,
    )


@app.route("/relatorios")
@login_required
def reports():
    coverage_rows = query_db(
        """
        SELECT COALESCE(NULLIF(TRIM(g.name), ''), 'Sem gestor') AS label, COUNT(*) AS total
        FROM bases b
        LEFT JOIN gestors g ON g.id = b.gestor_id
        GROUP BY label
        ORDER BY total DESC, label ASC
        """
    )
    coord_rows = query_db(
        """
        SELECT COALESCE(g.coordenacao, 'Sem coordenação') as label, COUNT(*) as total
        FROM bases b
        LEFT JOIN gestors g ON g.id = b.gestor_id
        GROUP BY label
        ORDER BY total DESC, label ASC
        """
    )
    env_rows = query_db(
        """
        SELECT COALESCE(NULLIF(TRIM(ambiente), ''), 'Sem ambiente') as label, COUNT(*) as total
        FROM bases
        GROUP BY label
        ORDER BY total DESC, label ASC
        """
    )

    data = {
        "total_gestors": query_db("SELECT COUNT(*) as total FROM gestors")[0]["total"],
        "total_bases": sum((row["total"] or 0) for row in coverage_rows),
        "coverage_labels": [row["label"] or "Sem gestor" for row in coverage_rows],
        "coverage_values": [row["total"] for row in coverage_rows],
        "coord_labels": [row["label"] or "Sem coordenação" for row in coord_rows],
        "coord_values": [row["total"] for row in coord_rows],
        "env_labels": [row["label"] or "Sem ambiente" for row in env_rows],
        "env_values": [row["total"] for row in env_rows],
    }

    return render_template("reports.html", **data)


def get_gestors(term=None):
    if term:
        like_term = f"%{term}%"
        return query_db(
            """
            SELECT * FROM gestors
            WHERE name LIKE ? OR secretaria LIKE ? OR coordenacao LIKE ? OR email LIKE ?
            ORDER BY name COLLATE NOCASE ASC
            """,
            (like_term, like_term, like_term, like_term),
        )
    return query_db("SELECT * FROM gestors ORDER BY name COLLATE NOCASE ASC")


def parse_gestor_id(raw_value):
    if not raw_value:
        return None
    candidate = raw_value.strip().split(" ", 1)[0]
    try:
        return int(candidate)
    except ValueError:
        return None


def gestor_id_by_name(name):
    if not name:
        return None
    result = query_db("SELECT id FROM gestors WHERE name = ? COLLATE NOCASE", (name.strip(),))
    if result:
        return result[0]["id"]
    return None


@app.route("/gestores")
@login_required
def list_gestors():
    search_term = request.args.get("q", "").strip()
    gestors = get_gestors(search_term)
    return render_template("gestors.html", gestors=gestors, query=search_term)


@app.route("/gestores/novo")
@login_required
def new_gestor_form():
    return render_template("gestor_form.html")


@app.route("/gestores/criar", methods=["POST"])
@login_required
def add_gestor():
    name = request.form.get("name", "").strip()
    secretaria = request.form.get("secretaria", "").strip()
    coordenacao = request.form.get("coordenacao", "").strip()
    email = request.form.get("email", "").strip()

    if not all([name, secretaria, coordenacao, email]):
        flash("Preencha todos os campos do gestor.", "error")
        return redirect(url_for("new_gestor_form"))

    execute_db(
        "INSERT INTO gestors (name, secretaria, coordenacao, email) VALUES (?, ?, ?, ?)",
        (name, secretaria, coordenacao, email),
    )
    flash("Gestor cadastrado com sucesso.", "success")
    return redirect(url_for("list_gestors"))


@app.route("/gestores/<int:gestor_id>/editar")
@login_required
def edit_gestor(gestor_id):
    gestor = query_db("SELECT * FROM gestors WHERE id = ?", (gestor_id,))
    if not gestor:
        flash("Gestor não encontrado.", "error")
        return redirect(url_for("list_gestors"))
    return render_template("gestor_form.html", gestor=gestor[0])


@app.route("/gestores/<int:gestor_id>/atualizar", methods=["POST"])
@login_required
def update_gestor(gestor_id):
    name = request.form.get("name", "").strip()
    secretaria = request.form.get("secretaria", "").strip()
    coordenacao = request.form.get("coordenacao", "").strip()
    email = request.form.get("email", "").strip()

    if not all([name, secretaria, coordenacao, email]):
        flash("Preencha todos os campos do gestor.", "error")
        return redirect(url_for("edit_gestor", gestor_id=gestor_id))

    execute_db(
        "UPDATE gestors SET name = ?, secretaria = ?, coordenacao = ?, email = ? WHERE id = ?",
        (name, secretaria, coordenacao, email, gestor_id),
    )
    flash("Gestor atualizado.", "success")
    return redirect(url_for("list_gestors"))


@app.route("/gestores/<int:gestor_id>/remover", methods=["POST"])
@login_required
def delete_gestor(gestor_id):
    in_use = query_db(
        "SELECT COUNT(*) as total FROM bases WHERE gestor_id = ? OR substituto1_id = ? OR substituto2_id = ?",
        (gestor_id, gestor_id, gestor_id),
    )[0]["total"]
    if in_use:
        flash("Não é possível remover: gestor vinculado a bases.", "error")
        return redirect(url_for("list_gestors"))

    execute_db("DELETE FROM gestors WHERE id = ?", (gestor_id,))
    flash("Gestor removido.", "success")
    return redirect(url_for("list_gestors"))


def gestor_choices(term=None):
    return [
        {
            "id": g["id"],
            "label": f"{g['id']} - {g['name']} ({g['secretaria']} / {g['coordenacao']})",
        }
        for g in get_gestors(term)
    ]


def selected_labels_from_base(base_row, options):
    label_map = {opt["id"]: opt["label"] for opt in options}
    return {
        "gestor": label_map.get(base_row["gestor_id"], ""),
        "sub1": label_map.get(base_row["substituto1_id"], ""),
        "sub2": label_map.get(base_row["substituto2_id"], ""),
    }


def ensure_gestor_exists(gestor_id, field_label):
    if not gestor_id:
        flash(f"Selecione um {field_label} válido a partir da lista.", "error")
        return False
    exists = query_db("SELECT id FROM gestors WHERE id = ?", (gestor_id,))
    if not exists:
        flash(f"{field_label} não encontrado.", "error")
        return False
    return True


@app.route("/bases")
@login_required
def list_bases():
    records = query_db(
        """
        SELECT b.*, g.name as gestor_name, gs1.name as sub1_name, gs2.name as sub2_name
        FROM bases b
        LEFT JOIN gestors g ON g.id = b.gestor_id
        LEFT JOIN gestors gs1 ON gs1.id = b.substituto1_id
        LEFT JOIN gestors gs2 ON gs2.id = b.substituto2_id
        ORDER BY b.id DESC
        """
    )
    return render_template("bases.html", bases=records)


@app.route("/bases/nova")
@login_required
def new_base_form():
    options = gestor_choices()
    if not options:
        flash("Cadastre pelo menos um gestor antes de criar bases.", "error")
        return redirect(url_for("new_gestor_form"))
    return render_template("base_form.html", gestors=options, selected_labels=None)


@app.route("/bases/criar", methods=["POST"])
@login_required
def add_base():
    name = request.form.get("name", "").strip()
    ambiente = request.form.get("ambiente", "").strip()
    descricao = request.form.get("descricao", "").strip()
    gestor_id = parse_gestor_id(request.form.get("gestor_id", ""))
    sub1_id = parse_gestor_id(request.form.get("substituto1_id", ""))
    sub2_id = parse_gestor_id(request.form.get("substituto2_id", ""))

    if not name:
        flash("Informe o nome da base.", "error")
        return redirect(url_for("new_base_form"))

    if not (gestor_id and ensure_gestor_exists(gestor_id, "Gestor")):
        return redirect(url_for("new_base_form"))

    if sub1_id and not ensure_gestor_exists(sub1_id, "1º substituto"):
        return redirect(url_for("new_base_form"))
    if sub2_id and not ensure_gestor_exists(sub2_id, "2º substituto"):
        return redirect(url_for("new_base_form"))

    if sub1_id and sub2_id and sub1_id == sub2_id:
        flash("Substitutos precisam ser pessoas diferentes.", "error")
        return redirect(url_for("new_base_form"))
    if gestor_id and (gestor_id == sub1_id or gestor_id == sub2_id):
        flash("Gestor titular não pode repetir um substituto.", "error")
        return redirect(url_for("new_base_form"))

    execute_db(
        """
        INSERT INTO bases (name, ambiente, descricao, gestor_id, substituto1_id, substituto2_id, source_connector)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            name,
            ambiente or None,
            descricao or None,
            gestor_id,
            sub1_id,
            sub2_id,
            "manual",
        ),
    )
    flash("Base cadastrada com sucesso.", "success")
    return redirect(url_for("list_bases"))


@app.route("/bases/<int:base_id>/editar")
@login_required
def edit_base(base_id):
    record = query_db("SELECT * FROM bases WHERE id = ?", (base_id,))
    if not record:
        flash("Base não encontrada.", "error")
        return redirect(url_for("list_bases"))
    options = gestor_choices()
    labels = selected_labels_from_base(record[0], options)
    return render_template("base_form.html", base=record[0], gestors=options, selected_labels=labels)


@app.route("/bases/<int:base_id>/atualizar", methods=["POST"])
@login_required
def update_base(base_id):
    name = request.form.get("name", "").strip()
    ambiente = request.form.get("ambiente", "").strip()
    descricao = request.form.get("descricao", "").strip()
    gestor_id = parse_gestor_id(request.form.get("gestor_id", ""))
    sub1_id = parse_gestor_id(request.form.get("substituto1_id", ""))
    sub2_id = parse_gestor_id(request.form.get("substituto2_id", ""))

    if not name:
        flash("Informe o nome da base.", "error")
        return redirect(url_for("edit_base", base_id=base_id))

    if not (gestor_id and ensure_gestor_exists(gestor_id, "Gestor")):
        return redirect(url_for("edit_base", base_id=base_id))

    if sub1_id and not ensure_gestor_exists(sub1_id, "1º substituto"):
        return redirect(url_for("edit_base", base_id=base_id))
    if sub2_id and not ensure_gestor_exists(sub2_id, "2º substituto"):
        return redirect(url_for("edit_base", base_id=base_id))

    if sub1_id and sub2_id and sub1_id == sub2_id:
        flash("Substitutos precisam ser pessoas diferentes.", "error")
        return redirect(url_for("edit_base", base_id=base_id))
    if gestor_id and (gestor_id == sub1_id or gestor_id == sub2_id):
        flash("Gestor titular não pode repetir um substituto.", "error")
        return redirect(url_for("edit_base", base_id=base_id))

    execute_db(
        """
        UPDATE bases
        SET name = ?, ambiente = ?, descricao = ?, gestor_id = ?, substituto1_id = ?, substituto2_id = ?
        WHERE id = ?
        """,
        (name, ambiente or None, descricao or None, gestor_id, sub1_id, sub2_id, base_id),
    )
    flash("Base atualizada.", "success")
    return redirect(url_for("list_bases"))


@app.route("/bases/<int:base_id>/remover", methods=["POST"])
@login_required
def delete_base(base_id):
    execute_db("DELETE FROM bases WHERE id = ?", (base_id,))
    flash("Base removida.", "success")
    return redirect(url_for("list_bases"))


@app.route("/buscar")
@login_required
def search():
    term = request.args.get("q", "").strip()
    tipo = request.args.get("tipo", "bases").strip()
    gestor = request.args.get("gestor", "").strip()
    base_nome = request.args.get("base", "").strip()
    ambiente = request.args.get("ambiente", "").strip()
    fonte = request.args.get("fonte", "").strip()
    descricao = request.args.get("descricao", "").strip()

    should_search = any([term, gestor, base_nome, ambiente, fonte, descricao])
    results = []

    if should_search and tipo == "bases":
        conditions = []
        params = []
        if term:
            like_term = f"%{term}%"
            conditions.append(
                "(b.name LIKE ? OR COALESCE(b.descricao, '') LIKE ? OR COALESCE(b.ambiente, '') LIKE ? OR COALESCE(g.name, '') LIKE ?)"
            )
            params.extend([like_term, like_term, like_term, like_term])
        if gestor:
            conditions.append("g.name LIKE ?")
            params.append(f"%{gestor}%")
        if base_nome:
            conditions.append("b.name LIKE ?")
            params.append(f"%{base_nome}%")
        if descricao:
            conditions.append("COALESCE(b.descricao, '') LIKE ?")
            params.append(f"%{descricao}%")
        if ambiente:
            conditions.append("b.ambiente = ?")
            params.append(ambiente)
        if fonte:
            conditions.append("b.source_connector = ?")
            params.append(fonte)

        query = """
            SELECT b.*, g.name as gestor_name
            FROM bases b
            LEFT JOIN gestors g ON g.id = b.gestor_id
        """
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
        query += " ORDER BY b.id DESC"
        results = query_db(query, tuple(params))

    ambientes = query_db(
        "SELECT DISTINCT ambiente FROM bases WHERE ambiente IS NOT NULL AND ambiente != '' ORDER BY ambiente"
    )
    fontes = query_db(
        "SELECT DISTINCT source_connector FROM bases WHERE source_connector IS NOT NULL AND source_connector != '' ORDER BY source_connector"
    )
    gestores = query_db(
        "SELECT DISTINCT name FROM gestors WHERE name IS NOT NULL AND name != '' ORDER BY name"
    )

    return render_template(
        "search.html",
        query=term,
        results=results,
        filters={
            "tipo": tipo,
            "gestor": gestor,
            "base": base_nome,
            "ambiente": ambiente,
            "fonte": fonte,
            "descricao": descricao,
        },
        options={
            "ambientes": [row["ambiente"] for row in ambientes],
            "fontes": [row["source_connector"] for row in fontes],
            "gestores": [row["name"] for row in gestores],
        },
    )


@app.route("/buscar/sugestoes")
@login_required
def search_suggestions():
    term = request.args.get("q", "").strip()
    suggestions = []
    if term:
        like = f"%{term}%"
        base_rows = query_db(
            "SELECT DISTINCT name FROM bases WHERE name LIKE ? ORDER BY name LIMIT 5", (like,)
        )
        gestor_rows = query_db(
            "SELECT DISTINCT name FROM gestors WHERE name LIKE ? ORDER BY name LIMIT 5", (like,)
        )
        ambiente_rows = query_db(
            "SELECT DISTINCT ambiente FROM bases WHERE ambiente LIKE ? AND ambiente IS NOT NULL ORDER BY ambiente LIMIT 5",
            (like,),
        )
        for row in base_rows:
            suggestions.append({"label": row["name"], "tipo": "Base"})
        for row in gestor_rows:
            suggestions.append({"label": row["name"], "tipo": "Gestor"})
        for row in ambiente_rows:
            suggestions.append({"label": row["ambiente"], "tipo": "Ambiente"})

    return jsonify({"suggestions": suggestions})


@app.route("/importar", methods=["GET", "POST"])
@login_required
def import_records():
    return render_template("import.html")


def require_import_data(flow):
    bucket = get_flow_bucket(flow)
    headers = bucket.get("headers") or []
    rows = bucket.get("rows") or []
    if not headers or not rows:
        flash("Envie um arquivo para começar o fluxo de importação.", "error")
        clear_import_state(flow)
        return None, None
    return headers, rows


@app.route("/importar/gestores", methods=["GET", "POST"])
@login_required
def import_gestors_flow():
    step = request.args.get("step", "upload")
    flow = "gestores"
    bucket = get_flow_bucket(flow)

    if request.method == "POST" and step == "upload":
        upload = request.files.get("file")
        delimiter = request.form.get("delimiter", ";").strip() or ";"

        if not upload or not upload.filename:
            flash("Selecione um arquivo CSV ou XLSX para continuar.", "error")
            return redirect(url_for("import_gestors_flow"))

        try:
            headers, rows = parse_tabular(upload, delimiter)
        except Exception:
            flash("Não foi possível ler o arquivo. Confirme o formato e o delimitador.", "error")
            return redirect(url_for("import_gestors_flow"))

        if not rows:
            flash("Nenhuma linha encontrada para importar.", "error")
            return redirect(url_for("import_gestors_flow"))

        bucket["headers"] = headers
        bucket["rows"] = rows
        bucket.pop("mapping", None)
        bucket.pop("result", None)
        return redirect(url_for("import_gestors_flow", step="mapear"))

    if request.method == "POST" and step == "mapear":
        headers, rows = require_import_data(flow)
        if headers is None:
            return redirect(url_for("import_gestors_flow"))

        mapping = {
            "name": request.form.get("map_name"),
            "secretaria": request.form.get("map_secretaria"),
            "coordenacao": request.form.get("map_coordenacao"),
            "email": request.form.get("map_email"),
        }

        if not all(mapping.values()):
            flash("Mapeie todas as colunas obrigatórias para seguir.", "error")
            return redirect(url_for("import_gestors_flow", step="mapear"))

        bucket["mapping"] = mapping
        return redirect(url_for("import_gestors_flow", step="confirmar"))

    if request.method == "POST" and step == "executar":
        headers, rows = require_import_data(flow)
        mapping = bucket.get("mapping")
        if headers is None or not mapping:
            return redirect(url_for("import_gestors_flow"))

        total = len(rows)
        imported = 0
        errors = []
        prepared = []

        header_set = set(headers)
        for row in rows:
            missing_cols = [col for col in mapping.values() if col not in header_set]
            if missing_cols:
                errors.append("Arquivo mudou: colunas mapeadas não foram encontradas.")
                break

            name = row.get(mapping["name"], "").strip()
            secretaria = row.get(mapping["secretaria"], "").strip()
            coordenacao = row.get(mapping["coordenacao"], "").strip()
            email = row.get(mapping["email"], "").strip()

            if not all([name, secretaria, coordenacao, email]):
                errors.append("Linha ignorada por falta de campos obrigatórios.")
                continue

            prepared.append((name, secretaria, coordenacao, email))

        if prepared:
            bulk_insert(prepared)
            imported = len(prepared)

        bucket["result"] = {
            "total": total,
            "imported": imported,
            "errors": errors,
            "progress": 100 if total else 0,
        }
        bucket["rows"] = prepared
        return redirect(url_for("import_gestors_flow", step="resultado"))

    if step == "mapear":
        headers, rows = require_import_data(flow)
        if headers is None:
            return redirect(url_for("import_gestors_flow"))
        return render_template("import_gestors.html", step="mapear", headers=headers, preview=rows[:5])

    if step == "confirmar":
        headers, rows = require_import_data(flow)
        mapping = bucket.get("mapping")
        if headers is None or not mapping:
            return redirect(url_for("import_gestors_flow"))

        preview = []
        for row in rows[:5]:
            preview.append(
                {
                    "name": row.get(mapping.get("name"), ""),
                    "secretaria": row.get(mapping.get("secretaria"), ""),
                    "coordenacao": row.get(mapping.get("coordenacao"), ""),
                    "email": row.get(mapping.get("email"), ""),
                }
            )

        return render_template(
            "import_gestors.html",
            step="confirmar",
            mapping=mapping,
            preview=preview,
            total=len(rows),
        )

    if step == "resultado":
        result = bucket.get("result")
        if not result:
            return redirect(url_for("import_gestors_flow"))
        return render_template("import_gestors.html", step="resultado", result=result)

    clear_import_state(flow)
    return render_template("import_gestors.html", step="upload")


@app.route("/importar/bases", methods=["GET", "POST"])
@login_required
def import_bases_flow():
    step = request.args.get("step", "upload")
    flow = "bases"
    bucket = get_flow_bucket(flow)

    if request.method == "POST" and step == "upload":
        upload = request.files.get("file")
        delimiter = request.form.get("delimiter", ";").strip() or ";"

        if not upload or not upload.filename:
            flash("Selecione um arquivo CSV ou XLSX para continuar.", "error")
            return redirect(url_for("import_bases_flow"))

        try:
            headers, rows = parse_tabular(upload, delimiter)
        except Exception:
            flash("Não foi possível ler o arquivo. Confirme o formato e o delimitador.", "error")
            return redirect(url_for("import_bases_flow"))

        if not rows:
            flash("Nenhuma linha encontrada para importar.", "error")
            return redirect(url_for("import_bases_flow"))

        bucket["headers"] = headers
        bucket["rows"] = rows
        bucket.pop("mapping", None)
        bucket.pop("result", None)
        return redirect(url_for("import_bases_flow", step="mapear"))

    if request.method == "POST" and step == "mapear":
        headers, rows = require_import_data(flow)
        if headers is None:
            return redirect(url_for("import_bases_flow"))

        mapping = {
            "name": request.form.get("map_name"),
            "ambiente": request.form.get("map_ambiente"),
            "descricao": request.form.get("map_descricao"),
            "gestor": request.form.get("map_gestor"),
            "sub1": request.form.get("map_sub1"),
            "sub2": request.form.get("map_sub2"),
        }

        required_fields = [mapping["name"], mapping["gestor"]]
        if not all(required_fields):
            flash("Mapeie ao menos Base e Gestor titular para continuar.", "error")
            return redirect(url_for("import_bases_flow", step="mapear"))

        bucket["mapping"] = mapping
        return redirect(url_for("import_bases_flow", step="confirmar"))

    if request.method == "POST" and step == "executar":
        headers, rows = require_import_data(flow)
        mapping = bucket.get("mapping")
        if headers is None or not mapping:
            return redirect(url_for("import_bases_flow"))

        total = len(rows)
        imported = 0
        errors = []
        prepared = []

        header_set = set(headers)
        required_map = [mapping["name"], mapping["gestor"]]

        for row in rows:
            missing_cols = [col for col in required_map if col not in header_set]
            optional_cols = [
                mapping.get("ambiente"),
                mapping.get("descricao"),
                mapping.get("sub1"),
                mapping.get("sub2"),
            ]
            missing_optional = [col for col in optional_cols if col and col not in header_set]
            if missing_cols or missing_optional:
                errors.append("Arquivo mudou: colunas mapeadas não foram encontradas.")
                break

            name = row.get(mapping["name"], "").strip()
            ambiente = row.get(mapping["ambiente"], "").strip() if mapping.get("ambiente") else ""
            descricao = row.get(mapping["descricao"], "").strip() if mapping.get("descricao") else ""
            gestor_name = row.get(mapping["gestor"], "").strip()
            sub1_name = row.get(mapping.get("sub1"), "").strip() if mapping.get("sub1") else ""
            sub2_name = row.get(mapping.get("sub2"), "").strip() if mapping.get("sub2") else ""

            if not all([name, gestor_name]):
                errors.append("Linha ignorada por falta de base ou gestor titular.")
                continue

            gestor_id = gestor_id_by_name(gestor_name)
            if not gestor_id:
                errors.append(f"Gestor '{gestor_name}' não encontrado.")
                continue

            sub1_id = None
            sub2_id = None

            if sub1_name:
                sub1_id = gestor_id_by_name(sub1_name)
                if not sub1_id:
                    errors.append(f"1º substituto '{sub1_name}' não encontrado.")
                    continue

            if sub2_name:
                sub2_id = gestor_id_by_name(sub2_name)
                if not sub2_id:
                    errors.append(f"2º substituto '{sub2_name}' não encontrado.")
                    continue

            if sub1_id and sub2_id and sub1_id == sub2_id:
                errors.append("Substitutos precisam ser pessoas diferentes.")
                continue
            if gestor_id and (gestor_id == sub1_id or gestor_id == sub2_id):
                errors.append("Gestor titular não pode repetir um substituto.")
                continue

            prepared.append((name, ambiente or None, descricao or None, gestor_id, sub1_id, sub2_id))

        if prepared:
            bulk_insert_bases(prepared)
            imported = len(prepared)

        bucket["result"] = {
            "total": total,
            "imported": imported,
            "errors": errors,
            "progress": 100 if total else 0,
        }
        bucket["rows"] = prepared
        return redirect(url_for("import_bases_flow", step="resultado"))

    if step == "mapear":
        headers, rows = require_import_data(flow)
        if headers is None:
            return redirect(url_for("import_bases_flow"))
        return render_template("import_bases.html", step="mapear", headers=headers, preview=rows[:5])

    if step == "confirmar":
        headers, rows = require_import_data(flow)
        mapping = bucket.get("mapping")
        if headers is None or not mapping:
            return redirect(url_for("import_bases_flow"))

        preview = []
        for row in rows[:5]:
            preview.append(
                {
                    "name": row.get(mapping.get("name"), ""),
                    "ambiente": row.get(mapping.get("ambiente"), "") if mapping.get("ambiente") else "",
                    "descricao": row.get(mapping.get("descricao"), "") if mapping.get("descricao") else "",
                    "gestor": row.get(mapping.get("gestor"), ""),
                    "sub1": row.get(mapping.get("sub1"), ""),
                    "sub2": row.get(mapping.get("sub2"), ""),
                }
            )

        return render_template(
            "import_bases.html",
            step="confirmar",
            mapping=mapping,
            preview=preview,
            total=len(rows),
        )

    if step == "resultado":
        result = bucket.get("result")
        if not result:
            return redirect(url_for("import_bases_flow"))
        return render_template("import_bases.html", step="resultado", result=result)

    clear_import_state(flow)
    return render_template("import_bases.html", step="upload")


@app.route("/extracao")
@login_required
def extraction_menu():
    return render_template("extract.html")


def prefill_from_job(job_id, bucket):
    job = get_job(job_id)
    if not job:
        return
    bucket["config"] = {
        "host": job["host"] or "",
        "jdbc_url": job["jdbc_url"] or "",
        "connection_type": job["connection_type"] or "TD2",
        "database_name": job["database_name"] or "",
        "username": job["username"] or "",
        "password": job["password"] or "",
        "extra_params": job["extra_params"] or "",
    }
    bucket["mode"] = job["mode"] or "incremental"
    bucket["extraction_type"] = job["extraction_type"] or "metadata"


@app.route("/extracao/teradata", methods=["GET", "POST"])
@login_required
def extract_teradata():
    step = request.args.get("step", "config")
    flow = "extracao_teradata"
    bucket = get_flow_bucket(flow)
    job_id_param = request.args.get("job_id")

    if job_id_param and not bucket.get("config"):
        try:
            prefill_from_job(int(job_id_param), bucket)
        except ValueError:
            pass

    if request.method == "POST" and step == "config":
        manual_jdbc = request.form.get("jdbc_url", "").strip()
        host = request.form.get("host", "").strip()
        database_name = request.form.get("database_name", "").strip()
        connection_type = request.form.get("connection_type", "TD2").strip() or "TD2"
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        extra_params = request.form.get("extra_params", "").strip()
        jdbc_url = manual_jdbc or build_jdbc_url(host, database_name, connection_type, extra_params)

        bucket["config"] = {
            "host": host,
            "jdbc_url": jdbc_url,
            "connection_type": connection_type,
            "database_name": database_name,
            "username": username,
            "password": password,
            "extra_params": extra_params,
        }

        if request.form.get("action") == "test":
            ok, message = test_teradata_connection(bucket["config"])
            flash(message, "success" if ok else "error")
            return redirect(url_for("extract_teradata", step="config", job_id=job_id_param))

        if not jdbc_url or not username or not password:
            flash("Preencha JDBC, usuário e senha para continuar.", "error")
            return redirect(url_for("extract_teradata", step="config", job_id=job_id_param))

        return redirect(url_for("extract_teradata", step="tipos"))

    if step == "tipos":
        if not bucket.get("config"):
            flash("Configure a conexão antes de selecionar o tipo.", "error")
            return redirect(url_for("extract_teradata"))

        if request.method == "POST":
            bucket["extraction_type"] = request.form.get("extraction_type", "metadata")
            bucket["mode"] = request.form.get("mode", "incremental")
            return redirect(url_for("extract_teradata", step="executar"))
        return render_template("extract_teradata.html", step="tipos", bucket=bucket)

    if step == "executar":
        if not bucket.get("config"):
            flash("Configure a conexão antes de executar.", "error")
            return redirect(url_for("extract_teradata"))

        if request.method == "POST":
            config = bucket.get("config", {})
            extraction_type = bucket.get("extraction_type", "metadata")
            mode = bucket.get("mode", "incremental")

            job_id = create_extraction_job("teradata", extraction_type, mode, config)
            result = run_teradata_job(config, mode, extraction_type, job_id)
            bucket["result"] = {"job_id": job_id, **result}
            flash("Extração finalizada.", "success" if not result["errors"] else "warning")
            return redirect(url_for("extract_teradata", step="executar"))

        return render_template("extract_teradata.html", step="executar", bucket=bucket)

    if step == "config":
        return render_template("extract_teradata.html", step="config", bucket=bucket)

    clear_import_state(flow)
    return render_template("extract_teradata.html", step="config", bucket=bucket)


@app.route("/configuracoes")
@login_required
def settings():
    users = query_db("SELECT id, username FROM users ORDER BY username ASC")
    return render_template("settings.html", users=users)


@app.route("/jobs")
@login_required
def monitor_jobs():
    jobs = query_db("SELECT * FROM extraction_jobs ORDER BY created_at DESC")
    return render_template("jobs.html", jobs=jobs)


@app.route("/jobs/<int:job_id>/restart", methods=["POST"])
@login_required
def restart_job(job_id):
    job = get_job(job_id)
    if not job:
        flash("Job não encontrado.", "error")
        return redirect(url_for("monitor_jobs"))

    config = {
        "host": job["host"],
        "jdbc_url": job["jdbc_url"],
        "connection_type": job["connection_type"],
        "database_name": job["database_name"],
        "username": job["username"],
        "password": job["password"],
        "extra_params": job["extra_params"],
    }
    update_extraction_job(job_id, status="pending", progress=0, error=None)
    result = run_teradata_job(config, job["mode"], job["extraction_type"], job_id)
    flash(
        "Job reiniciado.",
        "success" if not result.get("errors") else "warning",
    )
    return redirect(url_for("monitor_jobs"))


@app.route("/jobs/<int:job_id>/editar")
@login_required
def edit_job(job_id):
    job = get_job(job_id)
    if not job:
        flash("Job não encontrado.", "error")
        return redirect(url_for("monitor_jobs"))
    flash("Configuração carregada no fluxo de extração.", "info")
    return redirect(url_for("extract_teradata", job_id=job_id))


@app.route("/jobs/<int:job_id>/logs")
@login_required
def download_logs(job_id):
    job = get_job(job_id)
    if not job:
        flash("Job não encontrado.", "error")
        return redirect(url_for("monitor_jobs"))
    content = job["log"] or "Sem logs disponíveis."
    response = app.response_class(content, mimetype="text/plain")
    response.headers["Content-Disposition"] = f"attachment; filename=job_{job_id}_logs.txt"
    return response


@app.route("/usuarios/criar", methods=["POST"])
@login_required
def create_user():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")

    if not username or not password:
        flash("Preencha usuário e senha para adicionar.", "error")
        return redirect(url_for("settings"))

    try:
        execute_db(
            "INSERT INTO users (username, password) VALUES (?, ?)",
            (username, password),
        )
    except sqlite3.IntegrityError:
        flash("Nome de usuário já existe.", "error")
        return redirect(url_for("settings"))

    flash("Usuário criado com sucesso.", "success")
    return redirect(url_for("settings"))


@app.route("/usuarios/<int:user_id>/resetar", methods=["POST"])
@login_required
def reset_user(user_id):
    password = request.form.get("password", "")
    if not password:
        flash("Informe uma nova senha para continuar.", "error")
        return redirect(url_for("settings"))

    execute_db("UPDATE users SET password = ? WHERE id = ?", (password, user_id))
    flash("Senha atualizada.", "success")
    return redirect(url_for("settings"))


@app.route("/usuarios/<int:user_id>/remover", methods=["POST"])
@login_required
def delete_user(user_id):
    user = query_db("SELECT username FROM users WHERE id = ?", (user_id,))
    if not user:
        flash("Usuário não encontrado.", "error")
        return redirect(url_for("settings"))

    username = user[0]["username"]
    if username == session.get("user"):
        flash("Não é possível remover o usuário logado.", "error")
        return redirect(url_for("settings"))

    if username == ADMIN_USERNAME:
        flash("O usuário administrador padrão não pode ser removido.", "error")
        return redirect(url_for("settings"))

    execute_db("DELETE FROM users WHERE id = ?", (user_id,))
    flash("Usuário removido.", "success")
    return redirect(url_for("settings"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user"):
        return redirect(url_for("landing"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = query_db(
            "SELECT username FROM users WHERE username = ? AND password = ?",
            (username, password),
        )

        if user:
            session["user"] = username
            next_page = request.args.get("next") or url_for("landing")
            flash("Login realizado com sucesso.", "success")
            return redirect(next_page)

        flash("Credenciais inválidas.", "error")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("user", None)
    flash("Sessão encerrada.", "success")
    return redirect(url_for("login"))


init_db()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(debug=True, host="0.0.0.0", port=port)
